###################################################################
#
#  FEBID Simulation
#
#  Version 0.9
#
####################################################################
# Default packages
import datetime
import math
import os, sys
import random as rnd
import time
import warnings
import timeit
from threading import Thread
from tkinter import filedialog as fd

# Core packages
import line_profiler
import numpy as np
import pyvista as pv

# Auxillary packages
import pandas as pd
import openpyxl
# import matplotlib.pyplot as plt
# from matplotlib import cm
import yaml
from tqdm import tqdm

# Local packages
from source import Structure, Process
from source.libraries.vtk_rendering import VTK_Rendering as vr
from source.monte_carlo import etraj3d
import source.simple_patterns as sp

# It is assumed, that surface cell is a cell with a fully deposited cell(or substrate) under it and thus able produce deposit under irradiation.

# Semi-surface cells are cells that have precursor density but do not have a neighboring deposit cell
# Thus concept serves an alternative diffusion channel

flag = False
warnings.simplefilter('always')

class ThreadWithResult(Thread):
    def __init__(self, group=None, target=None, name=None, args=(), kwargs={}, *, daemon=None):
        self.result = None
        def function():
            self.result = target(*args, **kwargs)
        super().__init__(group=group, target=function, name=name, daemon=daemon)


class Statistics():
    """
    Class implementing statistics gathering and saving(to excel).

        Report contains following columns:

    Time, Time passed, Simulation time, Simulation speed, N of cells(filled), Growth speed, Simulation growth speed

        Additionally, initial simulation parameters are added to 3 separate sheets
    """

    def __init__(self, filename=f'run_id{rnd.randint(100000, 999999)}'):
        self.filename = filename
        self.columns = ['Time', 'Time passed', 'Sim.time', 'Sim.speed', 'N of cells', 'Growth speed', "Sim.growth rate"]
        self.units = ['', 's', 's', '', '', '1/s', '1/s']
        self.data = pd.DataFrame(columns=self.columns)
        self.data.loc[0] = [pd.Timestamp.now(), 0, 0, 0, 0, 0, 0]
        self.step = self.data.copy()
        self.parameters = []
        self.parameters_units = []

    def __getitem__(self, item):
        return self.data[item]

    @property
    def shape(self):
        return self.data.shape

    def get_params(self, arg: dict, name: str):
        """
        Collect initial parameters

        :param arg: a dictionary of parameters
        :param name: a name for the provided parameters
        :return:
        """
        series = pd.Series(arg)
        series.name = name
        self.parameters.append(series)

    def append(self, stats):
        """
        Add a new record to the statistics

        :param stats: current simulation time and current number of deposited cells
        :return:
        """
        self.dt = 0
        self.av_temperature = 0
        try:
            stats = (pd.Timestamp.now(), stats[0], stats[1])
            time_passed = (stats[0] - self.data.at[0, self.columns[0]]).total_seconds()
            sim_speed = stats[1] / time_passed
            growth_speed = stats[2] / time_passed * 60 * 60
            growth_rate = stats[2] / stats[1]
            # self.step = pd.Series({self.columns[1]:stats[0], self.columns[3]:stats[1]}, name=pd.Timestamp.now())
            # self.step.loc[self.shape[0]] = (stats[0], time_passed, stats[1], sim_speed, stats[2], growth_speed, growth_rate, stats[3])
            self.data.loc[self.shape[0]] = (stats[0], time_passed, stats[1], sim_speed, stats[2], growth_speed, growth_rate)
        except Exception as e:
            print(e.args)

        # self.data = self.data.append(self.step) # DataFrame.append() is not an in-place method like list.append()

    def plot(self, x, y):
        """
        ['Time', 'Sim.time', 'Sim.speed', 'N of cells', 'Growth rate', "Sim.growth rate"]
        :param x:
        :param y:
        :return:
        """
        if x not in self.columns or y not in self.columns:
            print(f'Column with this name does not exist!')
            return
        self._calculate_columns()
        self.plot(x=x, y=y)

    def save_to_file(self):
        """
        Write collected statistics to an excel file.

        If file does not exist, it is automatically created
        If file does exist, sheets with the names used here are overwritten
        """
        filename = self.filename + '.xlsx'
        vals = [c1 + ', ' + c2 for c1, c2 in zip(self.columns, self.units)]
        columns = zip(self.columns, vals)
        columns = dict(columns)
        data = self.data.copy()
        data.rename(columns=columns)
        sheet_name = 'Data'
        if not os.path.exists(filename):
            data.to_excel(filename, sheet_name=sheet_name, engine='openpyxl')
        else:
            try:
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    writer.book = openpyxl.load_workbook(filename)
                    if sheet_name in writer.book.sheetnames:
                        del writer.book[sheet_name]
                        wks = writer.book.create_sheet(sheet_name)
                        writer.sheets[sheet_name] = wks
                    data.to_excel(writer, sheet_name=sheet_name)
                    for params in self.parameters:
                        if params.name in writer.book.sheetnames:
                            del writer.book[params.name]
                            wks = writer.book.create_sheet(params.name)
                            writer.sheets[params.name] = wks
                        params.to_excel(writer, sheet_name=params.name)
            except Exception as e:
                print(e.args)
                sys.exit()

    def _calculate_columns(self):
        self.__get_time_passed()
        self.__get_sim_speed()
        self.__get_growth_speed()
        self.__get_sim_growth_rate()

    def __get_time_passed(self):
        self.data.iloc[1:, self.columns[1]] = self.data.loc[1:, 'Time'] - self.data.loc[:-1, 'Time']

    def __get_sim_speed(self):
        self.data.iloc['Sim.speed'] = self.data['Sim.time'] / self.data['Time']

    def __get_growth_speed(self):
        self.data['Growth rate'] = self.data['N of cells'] / self.data['Time'].get_total_hours

    def __get_sim_growth_rate(self):
        self.data['Sim.growth rate'] = self.data['N of cells'] / self.data['Sim.time'].get_total_hours


################################

def initialize_framework(from_file=False, precursor=None, settings=None, sim_params=None, vtk_file=None,
                         geom_params=None):
    """
        Open simulation configuration files and prepare data framework

    :param from_file: True to load structure from vtk file
    :param precursor: path to a file with precursor properties
    :param settings: path to a file with beam parameters and settings
    :param sim_params: path to a file with simulation volume parameters
    :param vtk_file: if from_file is True, path to a vtk file to get structure from
    :param geom_params: a list of predetermined simulation volume parameters
    :return:
    """

    print(f'Select three files: precursor parameters, beam parameters and either a predefined structure(.vtk) or a parameters file for the creation:')
    # precursor = None
    # settings = None
    # sim_params = None
    try:
        if precursor is None:
            precursor = yaml.load(open(fd.askopenfilename(), 'r'),
                                  Loader=yaml.Loader)  # Precursor and substrate properties(substrate here is the top layer)
        else:
            precursor = yaml.load(open(precursor, 'r'), Loader=yaml.Loader)
        if settings is None:
            settings = yaml.load(open(fd.askopenfilename(), 'r'),
                                 Loader=yaml.Loader)  # Parameters of the beam, dwell time and precursor flux
        else:
            settings = yaml.load(open(settings, 'r'), Loader=yaml.Loader)
        if not from_file:
            if sim_params is None and geom_params is None:
                sim_params = yaml.load(open(fd.askopenfilename(), 'r'),
                                       Loader=yaml.Loader)  # Size of the chamber, cell size and time step
            elif geom_params is not None:
                sim_params = dict()
                sim_params['width'] = geom_params[0]
                sim_params['length'] = geom_params[1]
                sim_params['height'] = geom_params[2]
                sim_params['cell_dimension'] = geom_params[4]
                sim_params['substrate_height'] = geom_params[3]
            else:
                sim_params = yaml.load(open(sim_params, 'r'), Loader=yaml.Loader)
        # precursor = yaml.load(open(f'{sys.path[0]}{os.sep}Me3PtCpMe.yml', 'r'), Loader=yaml.Loader)
        # settings = yaml.load(open(f'{sys.path[0]}{os.sep}Parameters.yml', 'r'),Loader=yaml.Loader)
        # sim_params = yaml.load(open(f'{sys.path[0]}{os.sep}Simulation.yml', 'r'),Loader=yaml.Loader)
    except Exception as e:
        print(e.args)
        sys.exit('An error occurred while reading one of the parameter files')

    structure = Structure()
    vtk_obj = None
    if from_file:
        try:
            if vtk_file is None:
                print(f'Specify .vtk structure file:')
                vtk_file = fd.askopenfilename()
            vtk_obj = pv.read(vtk_file)
            if not vtk_obj:
                raise RuntimeError
        except Exception as e:
            print(e.args)
            sys.exit("An error occurred while reading the VTK file")
        structure.load_from_vtk(vtk_obj=vtk_obj)
        sim_params = dict()
        sim_params['width'] = structure.shape[2] * structure.cell_dimension
        sim_params['length'] = structure.shape[1] * structure.cell_dimension
        sim_params['height'] = structure.shape[0] * structure.cell_dimension
        sim_params['cell_dimension'] = structure.cell_dimension
        sim_params['substrate_height'] = structure.substrate_height * structure.cell_dimension
    else:
        try:
            structure.create_from_parameters(sim_params['cell_dimension'], sim_params['width'], sim_params['length'],
                                             sim_params['height'], sim_params['substrate_height'])
        except:
            sys.exit("An error occurred while reading initial geometry parameters file")

    return structure, precursor, settings, sim_params


def buffer_constants(precursor: dict, settings: dict, cell_dimension: int):
    """
    Calculate necessary constants and prepare parameters for modules

    :param precursor: precursor properties
    :param settings: simulation conditions
    :param cell_dimension: side length of a square cell, nm
    :return:
    """

    elementary_charge = 1.60217662e-19
    # td = settings["dwell_time"]  # dwell time of a beam, s
    Ie = settings["beam_current"]  # beam current, A
    beam_FWHM = 2.36 * settings["gauss_dev"]  # electron beam diameter, nm
    F = settings[
        "precursor_flux"]  # precursor flux at the surface, 1/(nm^2*s)   here assumed a constant, but may be dependent on time and position
    effective_diameter = beam_FWHM * 3.3  # radius of an area which gets 99% of the electron beam
    f = Ie / elementary_charge / (
            math.pi * beam_FWHM * beam_FWHM / 4)  # electron flux at the surface, 1/(nm^2*s)
    e = precursor["SE_emission_activation_energy"]
    l = precursor["SE_mean_free_path"]

    # Precursor properties
    sigma = precursor[
        "cross_section"]  # dissociation cross section, nm^2; is averaged from cross sections of all electron types (PE,BSE, SE1, SE2)
    n0 = precursor["max_density"]  # inversed molecule size, Me3PtCpMe, 1/nm^2
    molar = precursor["molar_mass_precursor"]  # molar mass of the precursor Me3Pt(IV)CpMe, g/mole
    # density = 1.5E-20  # density of the precursor Me3Pt(IV)CpMe, g/nm^3
    V = precursor["dissociated_volume"]  # atomic volume of the deposited atom (Pt), nm^3
    D = precursor["diffusion_coefficient"]  # diffusion coefficient, nm^2/s
    tau = precursor["residence_time"] * 1E-6  # average residence time, s; may be dependent on temperature

    kd = F / n0 + 1 / tau + sigma * f  # depletion rate
    kr = F / n0 + 1 / tau  # replenishment rate
    nr = F / kr  # absolute density after long time
    nd = F / kd  # depleted absolute density
    t_out = 1 / (1 / tau + F / n0)  # effective residence time
    p_out = 2 * math.sqrt(D * t_out) / beam_FWHM

    # Initializing framework
    t_flux = 1 / (sigma + f)  # dissociation event time
    diffusion_dt = math.pow(cell_dimension * cell_dimension, 2) / (2 * D * (
            cell_dimension * cell_dimension + cell_dimension * cell_dimension))  # maximum stability
    dt = np.min([t_flux, diffusion_dt, tau])

    # Parameters for Monte-Carlo simulation
    mc_config = {'name': precursor["deposit"], 'E0': settings["beam_energy"],
                 'Emin': settings["minimum_energy"],
                 'Z': precursor["average_element_number"],
                 'A': precursor["average_element_mol_mass"], 'rho': precursor["average_density"],
                 'I0': settings["beam_current"], 'sigma': settings["gauss_dev"], 'n': settings['n'],
                 'N': Ie, 'sub': settings["substrate_element"],
                 'cell_dim': cell_dimension,
                 'e': precursor["SE_emission_activation_energy"], 'l': precursor["SE_mean_free_path"],
                  'emission_fraction': settings['emission_fraction']}
    # Parameters for reaction-equation solver
    equation_values = {'F': settings["precursor_flux"], 'n0': precursor["max_density"],
                       'sigma': precursor["cross_section"], 'tau': precursor["residence_time"] * 1E-6,
                       'V': precursor["dissociated_volume"], 'D': precursor["diffusion_coefficient"],
                       'dt': dt, 'deposition_scaling': settings['deposition_scaling']}
    # Stability time steps
    timings = {'t_diff': diffusion_dt, 't_flux': t_flux, 't_desorption': tau, 'dt': dt}

    # effective_radius_relative = math.floor(effective_diameter / cell_dimension / 2)
    return mc_config, equation_values, timings, nr


def start_from_command_line():
    """
    Simulation entry point from a command line
    :return:
    """
    from_file = input("Open predefined structure? [y/n] Otherwise a clear substrate will be created.")
    from_file = from_file.strip()
    if from_file not in ['y', 'n']:
        raise RuntimeError('Unacceptable input! Exiting.')
    elif from_file == 'y':
        from_file = True
    else:
        from_file = False
    structure, precursor_params, settings, sim_params = initialize_framework(from_file)
    print(f'Specify a stream file. If no file is specified, beam will be stationed in the center.')
    path = fd.askopenfilename()
    if not path:
        x, y = structure.shape[2] // 2, structure.shape[1] // 2
        path = sp.generate_pattern(path[0], 100000000, 1e-4, x, y, None)
    else:
        path, shape = sp.open_stream_file(path, True)
        structure.create_from_parameters(2, int(shape[2]) // 2, int(shape[1]) // 2, int(shape[0]) // 2, 4)

    run_febid(structure, precursor_params, settings, sim_params, path)


def run_febid_interface(structure, precursor_params, settings, sim_params, path, saving_params, rendering):

    if saving_params['monitoring']:
        dump_stats = True
        stats_rate = saving_params['monitoring']
    else:
        dump_stats = False
        stats_rate = math.inf
    if saving_params['snapshot']:
        dump_vtk = True
        dump_rate = saving_params['snapshot']
    else:
        dump_vtk = False
        dump_rate = math.inf

    kwargs = dict(location=saving_params['filename'], stats_rate=stats_rate,
                  dump_rate=dump_rate, render=rendering['show_process'],
                  frame_rate=rendering['frame_rate'], refresh_rate=0.5)
    run_febid(structure, precursor_params, settings, sim_params, path, dump_stats, kwargs)


def run_febid_test(geom, path, dwell_time, loops, files, kwargs):
    x, y = geom[0] // 2 * geom[4], geom[1] // 2 * geom[4]  # center
    path = sp.generate_pattern(path[0], loops, dwell_time, x, y, path[1:], step=2)

    structure, precursor_params, settings, sim_params = initialize_framework(False, files['precursor'],
                                                                             files['settings'], geom_params=geom)
    saving_params = {}
    run_febid(structure, precursor_params, settings, sim_params, path, True, kwargs)


def run_febid(structure, precursor_params, settings, sim_params, path, gather_stats=False, monitor_kwargs=None):
    """
        Create necessary objects and start the FEBID process

    :param structure: structure object
    :param precursor_params: precursor properties
    :param settings: beam and precursor flux settings
    :param sim_params: simulation volume properties
    :param path: printing path
    :param gather_stats: True enables statistics gathering
    :param monitor_kwargs: settings for the monitoring function
    :return:
    """
    mc_config, equation_values, timings, nr = buffer_constants(precursor_params, settings, sim_params['cell_dimension'])
    stats = None
    if gather_stats:
        stats = Statistics(monitor_kwargs['location'])
        stats.get_params(precursor_params, 'Precursor parameters')
        stats.get_params(settings, 'Beam parameters and settings')
        stats.get_params(sim_params, 'Simulation volume parameters')
    process_obj = Process(structure, equation_values, timings)
    sim = etraj3d.cache_params(mc_config, structure.deposit, structure.surface_bool, structure.surface_neighbors_bool)
    process_obj.max_neib = math.ceil(np.max([sim.deponat.lambda_escape, sim.substrate.lambda_escape])/process_obj.cell_dimension)
    process_obj.structure.define_surface_neighbors(process_obj.max_neib)
    total_iters = int(np.sum(path[:, 2]) / process_obj.dt)
    # Actual simulation runs in a second Thread, because visualization of the process
    # via Pyvista works only from the main Thread
    printing = Thread(target=print_all, args=[path, process_obj, sim])
    printing.start()
    monitoring(process_obj, total_iters, stats, **monitor_kwargs)
    printing.join()
    print('Finished path.')

def print_all(path, process_obj, sim):
    global flag
    start = 0
    av_dwell_time = path[:, 2].mean()
    # av_loops = int(path.shape[0] * av_dwell_time / process_obj.dt)
    av_loops = path[:,2].sum() / process_obj.dt
    total_time = path[:,2].sum()
    t = tqdm(total=av_loops, desc='total', position=0)
    for x, y, step in path[start:]:
        process_obj.beam_matrix[:, :, :] = etraj3d.rerun_simulation(y, x, sim, process_obj.dt)
        if process_obj.beam_matrix.max() <= 1:
            warnings.warn('No surface flux!', RuntimeWarning)
            process_obj.beam_matrix[...] = 1
        process_obj.get_dt()
        process_obj.update_helper_arrays()
        av_loops = path[:, 2].sum() / process_obj.dt
        t.total = av_loops
        print_step(y, x, step, process_obj, sim, t)
    flag = True


def print_step(y, x, dwell_time, pr: Process, sim, t):
    """
    Run deposition on a single spot

    :param x: spot x-coordinate
    :param y: spot y-coordinate
    :param dwell_time: time of the exposure
    :param pr: Process object
    :param sim: MC simulation object
    :return:
    """
    loops = int(dwell_time / pr.dt)
    time = 0
    while time < dwell_time:  # loop repeats
        pr.deposition()  # depositing on a selected area
        if pr.check_cells_filled():
            flag = pr.update_surface()  # updating surface on a selected area
            if flag:
                sim.grid = sim.m3d.grid = pr.deposit
                sim.surface = sim.m3d.surface = pr.surface
                sim.s_neighb = sim.m3d.s_neighb = pr.surface_n_neighbors
            start = timeit.default_timer()
            pr.beam_matrix[...] = etraj3d.rerun_simulation(y, x, sim, pr.dt)
            print(f'Finished MC in {timeit.default_timer()-start} s')
            if pr.beam_matrix.max() <= 1:
                warnings.warn('No surface flux!', RuntimeWarning)
                pr.beam_matrix[...] = 1
                continue
            pr.get_dt()
            pr.update_helper_arrays()
            a = 1
        pr.precursor_density()
        pr.t += pr.dt
        time += pr.dt
        t.update(1)


def monitoring(pr: Process, l, stats: Statistics = None, location=None, stats_rate=60, dump_rate=60, render=False,
               frame_rate=1, refresh_rate=0.5):
    """
    A daemon process function to manage statistics gathering and graphics update

    :param pr: object of the core deposition process
    :param l: approximate number of iterations
    :param stats: object for gathering monitoring data
    :param location: file saving directory
    :param stats_rate: statistics recording interval in seconds, None disables statistics recording
    :param dump_rate: dumping interval in seconds, None disables structure dumping
    :param render: True will enable graphical monitoring of the process
    :param frame_rate: redrawing delay
    :param refresh_rate: sleep time
    :return:
    """

    global flag  # This flag variable is used for the communication between current and the process thread.
    # When deposition process thread finishes, it sets flag to False which will finish current thread

    time_step = 60
    pr.start_time = datetime.datetime.now()
    dump_time = stats_time = 0
    time_spent = start_time = frame = timeit.default_timer()
    # Initializing graphical monitoring
    rn = None
    if render:
        rn = vr.Render(pr.structure.cell_dimension)
        pr.redraw = True
    else:
        frame = np.inf  # current time is always less than infinity
    if not stats_rate or stats_rate == np.inf:
        stats_time = np.inf
        stats_rate = np.inf
    else:
        stats_rate = 1e-1 * stats_rate
    if not dump_rate or dump_rate == np.inf:
        dump_time = np.inf
        dump_rate = np.inf
    else:
        dump_rate = 1e-1 * dump_rate
    # Event loop
    with open('monitor.log', mode='a') as f:
        f.write(f'Beginning monitor log, {time.strftime("%H:%M:%S", time.localtime())}:\n')
        while not flag:
            now = timeit.default_timer()
            if now > time_spent:  # overall time and speed
                time_spent += time_step
                # print(f'Time passed: {time_spent}, Av.speed: {l / time_spent}')
            if now > frame:  # graphical
                frame += frame_rate
                redrawed = update_graphical(rn, pr, frame_rate, now - start_time)
                if redrawed:
                    f.write(f'[{time.strftime("%H:%M:%S", time.localtime())}] Redrawed scene.\n')
            if pr.t > stats_time:
                stats_time += stats_rate
                stats.append((pr.t, np.count_nonzero(pr.deposit == -1)))
                stats.save_to_file()
                f.write(f'[{time.strftime("%H:%M:%S", time.localtime())}] Recorded statistics.\n')
            if pr.t > dump_time:
                dump_time += dump_rate
                dump_structure(pr.structure, f'{location}')
                f.write(f'[{time.strftime("%H:%M:%S", time.localtime())}] Dumped structure.\n')
            time.sleep(refresh_rate)
        else:
            if stats_time != np.inf:
                stats.append((pr.t, np.count_nonzero(pr.deposit == -1)))
                stats.save_to_file()
                f.write(f'[{time.strftime("%H:%M:%S", time.localtime())}] Recorded statistics.\n')
            if dump_time != np.inf:
                dump_structure(pr.structure, f'{location}')
                f.write(f'[{time.strftime("%H:%M:%S", time.localtime())}] Dumped structure.\n')
            if frame != np.inf:
                rn.p.close()
                rn = vr.Render(pr.structure.cell_dimension)
                pr.redraw = True
                update_graphical(rn, pr, time_step, now-start_time, False)
                rn.show(interactive_update=False)
                f.write(f'[{time.strftime("%H:%M:%S", time.localtime())}] Redrawed scene.\n')
    flag = False
    print('Exiting monitoring.')


def update_graphical(rn: vr.Render, pr: Process, time_step, time_spent, update=True):
    """
    Update the visual representation of the current process state

    :param rn: visual scene object
    :param pr: process object
    :param time_step:
    :param time_spent:
    :return:
    """
    data = pr.precursor
    redrawed = pr.redraw
    if pr.redraw:
        try:
            rn.y_pos = 5
            try:
                rn.p.button_widgets.clear()
            except: pass
            rn.p.clear()
            current = 'precursor'
            rn._add_3Darray(data, opacity=1, scalar_name='Precursor',
                            button_name='precursor', show_edges=True, cmap='plasma')
            scalar = rn.p.mesh.active_scalars_name
            rn.p.mesh[scalar] = data.reshape(-1)
            rn.update_mask(pr.surface)
            rn.p.add_text('.', position='upper_left', font_size=12, name='time')
            rn.p.add_text('.', position='upper_right', font_size=12, name='stats')
            rn.show(interactive_update=True, cam_pos=[(206.34055818793468, 197.6510638707941, 100.47106597548205),
                                                      (0.0, 0.0, 0.0),
                                                      (-0.23307751464125356, -0.236197909312718, 0.9433373838690787)])
        except Exception as e:
            print('An error occurred while redrawing the scene.')
            print(e.args)
            pass
        rn.meshes_count += 1
        pr.redraw = False
    # Calculating values to indicate
    pr.n_filled_cells.append(pr.filled_cells)
    i = len(pr.n_filled_cells) - 1
    time_real = str(datetime.timedelta(seconds=int(time_spent)))
    speed = pr.t / time_spent
    height = (pr.max_z - pr.substrate_height) * pr.structure.cell_dimension
    total_V = int((pr.filled_cells + pr.deposit[pr.surface].sum()) * pr.cell_V)
    growth_rate = int((total_V-pr.vol_prev) / (pr.t-pr.t_prev) / pr.deposition_scaling)
    pr.t_prev = pr.t
    pr.vol_prev = total_V
    # Updating displayed text
    rn.p.textActor.renderer.actors['time'].SetText(2,
                    f'Time: {time_real} \n' # showing real time passed 
                    f'Sim. time: {(pr.t*pr.deposition_scaling):.8f} s \n' # showing simulation time passed
                    f'Speed: {speed:.8f} \n'  
                    f'Av. growth rate: {growth_rate} nm^3/s')
    rn.p.textActor.renderer.actors['stats'].SetText(3,
                    f'Cells: {pr.n_filled_cells[i]} \n'  # showing total number of deposited cells
                    f'Height: {height} nm \n'
                    f'Volume: {total_V} nm^3')
    # Updating scene
    rn.update_mask(pr.surface)
    try:
        min = data[data > 0.00001].min()
    except:
        min = 1e-8
    rn.p.update_scalar_bar_range(clim=[min, data.max()])

    if update:
        rn.update()
    return redrawed


def dump_structure(structure: Structure, filename='FEBID_result'):
    vr.save_deposited_structure(structure, filename)


if __name__ == '__main__':
    print(f'##################### FEBID Simulator ###################### \n\n'
          f'Two modes are available: deposition process and Monte Carlo electron beam-matter simulation\n'
          f'Type \'febid\' to enter deposition mode or \'mc\' for electron trajectory simulation:')
    mode_choice = input("Enter 'febid', 'mc' or press Enter to exit:")
    if mode_choice not in ['febid', 'mc']:
        sys.exit("Exit program.")
    if mode_choice == 'mc':
        print(f'Entering Monte Carlo electron beam-matter simulation module')
        etraj3d.mc_simulation()
    if mode_choice == 'febid':
        start_from_command_line()
